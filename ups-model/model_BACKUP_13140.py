# -*- coding: utf-8 -*-
"""
Created on Tue Apr  3 21:08:04 2018

@author: Tzu-Ying, Pear, Yi
Project name: UPS Optimization Model for Network Design
"""

import pulp
import math
import openpyxl

data = openpyxl.load_workbook('list_zipcode.xlsx',read_only=True, data_only=True)
sheet = data['Sheet1']
title = []
for row in sheet.rows:
    
    for cell in row:
        title.append(cell.value)


data = openpyxl.load_workbook('cost_matrix_truck.xlsx',read_only=True, data_only=True)
sheet = data['cost']

cost_truck = {}

#inputting cost for the truck 
# structure would be like {ori:{dest: ***, dest2:***} ori2: {dest:...}}
for row in sheet.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    ori = temp[0]
    temp.pop(0)
    cost_truck[ori] = {}
    for i in range(len(title)):
        cost_truck[ori][title[i]] = []
        cost_truck[ori][title[i]].append(temp[i])



data2 = openpyxl.load_workbook('cost_matrix_deliveryday.xlsx', read_only=True, data_only=True)
sheet2 = data2['day']

delivery_day = {}

<<<<<<< HEAD
#inputting require days for each ori to dest
# structure would be like {ori:{dest: ***, dest2:***} ori2: {dest:...}}
=======
>>>>>>> origin/pear

for row in sheet2.rows:
    temp = []

    for cell in row:
        temp.append(cell.value)
    ori = temp[0]
    temp.pop(0)
    delivery_day[ori] = {}
<<<<<<< HEAD
    for i in range(len(title)):
        delivery_day[ori][title[i]] = []
        delivery_day[ori][title[i]].append(temp[i])



# creating dictionary for next day air cost from ori to dest
# structure would be like {ori:{dest: ***, dest2:***} ori2: {dest:...}}
data = openpyxl.load_workbook('cost_matrix_nda.xlsx',read_only=True, data_only=True)
sheet = data['cost']

cost_nda = {}

for row in sheet.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    ori = temp[0]
    temp.pop(0)
    cost_nda[ori] = {}
    for i in range(len(title)):
        cost_nda[ori][title[i]] = []
        cost_nda[ori][title[i]].append(temp[i])


# creating dictionary for 2nd day air cost from ori to dest
# structure would be like {ori:{dest: ***, dest2:***} ori2: {dest:...}}
data = openpyxl.load_workbook('cost_matrix_sda.xlsx',read_only=True, data_only=True)
sheet = data['cost']

cost_sda = {}

for row in sheet.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    ori = temp[0]
    temp.pop(0)
    cost_sda[ori] = {}
    for i in range(len(title)):
        cost_sda[ori][title[i]] = []
        cost_sda[ori][title[i]].append(temp[i])

# creating dictionary for demand of high value and value of each 3 digit
# structure would be like {3digit: [high, low], ...}
data = openpyxl.load_workbook('demand.xlsx',read_only=True, data_only=True)
sheet = data['Demand']

demand = {}

for row in sheet.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    zipcode = temp[0]
    temp.pop(0)
    demand[zipcode] = []
    for i in range(len(temp)):
        demand[zipcode].append(temp[i])

# creating dictionary for demand of high value and value of each 3 digit
# structure would be like {3digit: cost, 3digit2: cost, ... }
data = openpyxl.load_workbook('facility_cost.xlsx',read_only=True, data_only=True)
sheet = data['Demand']

facility_cost = {}

for row in sheet.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    zipcode = temp[0]
    temp.pop(0)
    facility_cost[zipcode] = temp[0]

=======
    for i in title:
        delivery_day[ori][i] = temp[i]
        
        
        
 #pear optimization code

#parameters
 #i = customer node i (clustered into 3-digits)
 #j = facility 
 #m = mode type [Ah,Al,Th,Tl]
     #Ah = air for high value goods
     #Al = air for low value goods
     #Th = truck for HV goods
     #Tl = truck for LV goods
 
 
#data section
  
# Z: a set of customer locations
#F: set of 3-digit zip code
#d[i,j] :distance i → j
#u[i,j] : binary. 1 if i  can be assigned to j at all; 0 otherwise.
#F = fixed cost to operate
#i: customer
#j: air hub/DC
#C: cost
     #C[i,j,Ah] = cost per mile per unit weight of HV goods via air from i to j
     #C[i,j,Al] = cost/mile/unit of LV goods via air from i to j
     #
#t[i,j,m] = either Aij or Bij

#decision variables
#X[i,j,m]  =  binary. 1 if i assigned to DC j served by mode m; 0 otherwise
#Y[j]  = binary. If facility j is open; 0 otherwise

 #objective function
 #   T*C[i,j]*X[i,j] + AC[i,j]*(1-F)+F*y[j]+f[j]
 
 #constraint 1 
 #for all j: sum(y[j]) = p(# of facilities to locate)
 
 #constraint 2 
 #for all i:
 #sum(across j sum(across m(x[i,j,m]))) >=1
 
 #constraint 3
 #for all j:
 #sum(across n sum(across i (x[i,j,m]))) <= y[j]
 
 #constraint 4:
 #for all j:
 #x[1,c]+x[2,c]+x[3,c] <= Y[j]
 
 
 
 
 
 
>>>>>>> origin/pear
