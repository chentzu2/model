# -*- coding: utf-8 -*-
"""
Created on Tue Apr  3 21:08:04 2018

@author: Tzu-Ying, Pear, Yi
Project name: UPS Optimization Model for Network Design
"""

import pulp
import math
import openpyxl

data = openpyxl.load_workbook('list_zipcode.xlsx',read_only=True, data_only=True)
sheet = data['Sheet1']
title = []
for row in sheet.rows:
    
    for cell in row:
        title.append(cell.value)


data = openpyxl.load_workbook('cost_matrix_truck.xlsx',read_only=True, data_only=True)
sheet = data['cost']

cost_truck = {}

#inputting cost for the truck 
# structure would be like {ori:{dest: ***, dest2:***} ori2: {dest:...}}
for row in sheet.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    ori = temp[0]
    temp.pop(0)
    cost_truck[ori] = {}
    for i in range(len(title)):
        cost_truck[ori][title[i]] = []
        cost_truck[ori][title[i]].append(temp[i])



data2 = openpyxl.load_workbook('cost_matrix_deliveryday.xlsx', read_only=True, data_only=True)
sheet2 = data2['day']

delivery_day = {}

#inputting require days for each ori to dest
# structure would be like {ori:{dest: ***, dest2:***} ori2: {dest:...}}

for row in sheet2.rows:
    temp = []

    for cell in row:
        temp.append(cell.value)
    ori = temp[0]
    temp.pop(0)
    delivery_day[ori] = {}
    for i in range(len(title)):
        delivery_day[ori][title[i]] = []
        delivery_day[ori][title[i]].append(temp[i])



# creating dictionary for next day air cost from ori to dest
# structure would be like {ori:{dest: ***, dest2:***} ori2: {dest:...}}
data = openpyxl.load_workbook('cost_matrix_nda.xlsx',read_only=True, data_only=True)
sheet = data['cost']

cost_nda = {}

for row in sheet.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    ori = temp[0]
    temp.pop(0)
    cost_nda[ori] = {}
    for i in range(len(title)):
        cost_nda[ori][title[i]] = []
        cost_nda[ori][title[i]].append(temp[i])


# creating dictionary for 2nd day air cost from ori to dest
# structure would be like {ori:{dest: ***, dest2:***} ori2: {dest:...}}
data = openpyxl.load_workbook('cost_matrix_sda.xlsx',read_only=True, data_only=True)
sheet = data['cost']

cost_sda = {}

for row in sheet.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    ori = temp[0]
    temp.pop(0)
    cost_sda[ori] = {}
    for i in range(len(title)):
        cost_sda[ori][title[i]] = []
        cost_sda[ori][title[i]].append(temp[i])

# creating dictionary for demand of high value and value of each 3 digit
# structure would be like {3digit: [high, low], ...}
data = openpyxl.load_workbook('demand.xlsx',read_only=True, data_only=True)
sheet = data['Demand']

demand = {}

for row in sheet.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    zipcode = temp[0]
    temp.pop(0)
    demand[zipcode] = []
    for i in range(len(temp)):
        demand[zipcode].append(temp[i])

# creating dictionary for demand of high value and value of each 3 digit
# structure would be like {3digit: cost, 3digit2: cost, ... }
data = openpyxl.load_workbook('facility_cost.xlsx',read_only=True, data_only=True)
sheet = data['Demand']

facility_cost = {}

for row in sheet.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    zipcode = temp[0]
    temp.pop(0)
    facility_cost[zipcode] = temp[0]

