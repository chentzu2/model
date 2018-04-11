# -*- coding: utf-8 -*-
"""
Created on Tue Apr  3 21:08:04 2018

@author: Tzu-Ying, Pear, Yi
Project name: UPS Optimization Model for Network Design
"""

#testing github desktop

import pulp
import math
import openpyxl

data = openpyxl.load_workbook('list_zipcode.xlsx',read_only=True, data_only=True)
sheet = data['Sheet1']
title = []
for row in sheet.rows:
    
    for cell in row:
        title.append(cell.value)


data = openpyxl.load_workbook('cost_matrix_truck.xlsx',read_only=True, data_only=True)
sheet = data['cost']

cost_truck = {}

#inputting cost for the truck {i1:{j1:C[i1,j1,m],j2:C[i1,j2,m]...}, i2:{j1:C[i2,j1,m], j2:C[i2,j2,m]...}....} (m=T)
# structure would be like {ori:{dest: ***, dest2:***} ori2: {dest:...}}
for row in sheet.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    ori = temp[0]
    temp.pop(0)
    cost_truck[ori] = {}
    for i in range(len(title)):
        cost_truck[ori][title[i]] = []
        cost_truck[ori][title[i]].append(temp[i])



data2 = openpyxl.load_workbook('cost_matrix_deliveryday.xlsx', read_only=True, data_only=True)
sheet2 = data2['day']

delivery_day = {}

#inputting require days for each ori to dest (truck transport)
# structure would be like {ori:{dest: ***, dest2:***} ori2: {dest:...}}

for row in sheet2.rows:
    temp = []

    for cell in row:
        temp.append(cell.value)
    ori = temp[0]
    temp.pop(0)
    delivery_day[ori] = {}
    for i in range(len(title)):
        delivery_day[ori][title[i]] = []
        delivery_day[ori][title[i]].append(temp[i])



# creating dictionary for next day air cost from ori to dest
# structure would be like {ori:{dest: ***, dest2:***} ori2: {dest:...}}
data = openpyxl.load_workbook('cost_matrix_nda.xlsx',read_only=True, data_only=True)
sheet = data['cost']

cost_nda = {}

for row in sheet.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    ori = temp[0]
    temp.pop(0)
    cost_nda[ori] = {}
    for i in range(len(title)):
        cost_nda[ori][title[i]] = []
        cost_nda[ori][title[i]].append(temp[i])


# creating dictionary for 2nd day air cost from ori to dest
# structure would be like {ori:{dest: ***, dest2:***} ori2: {dest:...}}
data = openpyxl.load_workbook('cost_matrix_sda.xlsx',read_only=True, data_only=True)
sheet = data['cost']

cost_sda = {}

for row in sheet.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    ori = temp[0]
    temp.pop(0)
    cost_sda[ori] = {}
    for i in range(len(title)):
        cost_sda[ori][title[i]] = []
        cost_sda[ori][title[i]].append(temp[i])

# creating dictionary for demand of high value and value of each 3 digit
# structure would be like {3digit: [high, low], ...}
data = openpyxl.load_workbook('demand.xlsx',read_only=True, data_only=True)
sheet = data['Demand']

demand = {}

for row in sheet.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    zipcode = temp[0]
    temp.pop(0)
    demand[zipcode] = []
    for i in range(len(temp)):
        demand[zipcode].append(temp[i])

# creating dictionary for facility F
# structure would be like {3digit: cost, 3digit2: cost, ... }
data = openpyxl.load_workbook('facility_cost.xlsx',read_only=True, data_only=True)
sheet = data['Demand']

facility_cost = {}

for row in sheet.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    zipcode = temp[0]
    temp.pop(0)
    facility_cost[zipcode] = temp[0]
    

#pear optimization code

# Data Section
CustLoc=[] #a set of customer locations
Zip=[] # set of 3-digit zip code
shipmax = 1 #change from next day to 2 days - connect to GUI
F = 10000 #fixed cost to operate/open facility  
<<<<<<< HEAD
travelMode = [1,2] #[air, truck]
M=10000000
costAir={1:cost_nda, 2:cust_sda} #{ori:{dest: ***, dest2:***} ori2: {dest:...}}
trucktime= delivery_day #{} #delivery_day={ori:{dest: ***, dest2:***} ori2: {dest:...}}
costTruck = cost_Truck #{ori:{dest: ***, dest2:***} ori2: {dest:...}}
d = demand #{3digit: [high, low], ...}
=======
V = [20,150]#V[k] = [Vl, Vh] - value of LV and HV goods = [20,150]
M = [1,2] #[air, truck]
cost_air={1:cost_nda, 2:cost_sda}

>>>>>>> master
# Code Section

# Create the 'prob' object to contain the problem data
 #create prob object to contain optimization problem data
problem = pulp.LpProblem("Facility Location Plan", pulp.LpMinimize)

# Decision variables

<<<<<<< HEAD
combo={} #={(1,2):{'route}}
for i,j in zip(Zip, CustLoc):
	combo[(i,j)]:{} #origin, destination 

for a,a_dict in combo.iteritems():
	z = a[0]
	cl = a[1]

	varAH = pulp.LpVariable("AirHigh(%s,%s)"%(str(z),str(cl)), lowBound=0, cat='Binary')
	varAL = pulp.LpVariable("AirLow(%s,%s)"%(str(z),str(cl)), lowBound=0, cat='Binary')
	varTH = pulp.LpVariable("TruckHigh(%s,%s)"%(str(z),str(cl)), lowBound=0, cat='Binary')
	varTL = pulp.LpVariable("TruckLow(%s,%s)"%(str(z),str(cl)), lowBound=0, cat='Binary')

	# a_dict['dvAH'] = varAH
	# a_dict['dvAL'] = varAL
	# a_dict['dvTH'] = varTH
	# a_dict['dvTL'] = varTL	
	#value = [l,h]
	a_dict['dv']=[varAH, varAL, varTH, varTL]
	objFn.append((varAH*d[cl][0]+varAL*d[cl][1])*costAir[shipmax][z][cl]*150 + (varTH*D[cl][0]+varAL*D[cl][1])*costTruck[z][cl]*20)


FacilityLocations={}
for j in Zip:
	dvLoc = pulp.LpVariable("Zip(%s)"%str(j), lowBound=0, upBound=1, cat='Binary')
	FacilityLocations[j] = dvLoc
	objFn.append(dvLoc*F)

prob += pulp.lpSum(objFn), "Total Cost"

=======
x = pulp.LpVariable.dicts("X", (Zip, CustLoc, M, V),  cat=pulp.LpBinary) #x[i,j,m,k]  =  binary. 1 if i assigned to DC j served by mode m for k value goods; 0 otherwise
y = pulp.LpVariable.dicts('Y', Zip,  cat = pulp.LpBinary) #y[j]  = binary. If facility j is open; 0 otherwise

# Now the ship variables
combo = pulp.LpVariable.dicts("route", (Zip, CustLoc), cat=pulp.LpBinary) #??
tr = pulp.LpVariable.dicts("truckTravelTime",Zip,lowBound=0) #t[i,j,m] t={1:shipmax, 2:{delivery_day}} 1 = air, 2 = truck
#delivery_day={ori:{dest: ***, dest2:***} ori2: {dest:...}}
C = pulp.LpVariable.dicts("Cost",M,lowBound=0) #C[m[i,j]] C={1:cost_air,2:cost_truck} 1 = air, 2= truck
d = pulp.LpVariable.dicts("Demand",Zip, lowBound =0) #d = demand // D[j,k]: demand of good type k at j {3digit: [high, low], ...}
t = {1: shipmax, 2: tr}
# Objective function
# The objective function is always added to 'prob' first in PuLP
#LpAfflineExpression

#!!!!! fix objective function, see below
# for a,a_dict in arcs.iteritems():
# 	i = a[0]
# 	j = a[1]
	
# 	# First create a total integer trailer flow variable, and tie it to the arc
# 	var = pulp.LpVariable("TrailerFlow(%s,%s)" % (str(i),str(j)), lowBound = 0, cat=pulp.LpInteger)
# 	a_dict['dvTrailerFlow'] = var
	
# 	# Add objective function term to objFn list variable
# 	if a_dict['cost'] != 0 :
# 		objFn.append(a_dict['cost']*var)
	
# 	# Create an empty dictionary inside the arc attributes dictionary
# 	# to hold the fractional flow decision variables for each commodity
# 	a_dict['dvFlows']={}
# 	# Add a decision variable for each commodity k
# 	for k in commods:
# 		orig = k[0]
# 		dest = k[1]
# 		# Format for LpVariable("Name",lowBound,cat)
# 		# Name will list the arc first, then the commodity; e.g.
# 		# Arc_Flow('ATL,'DAL')_('Ath','CHI')
# 		var = pulp.LpVariable("ArcFlow(%s,%s)_(%s,%s)" % (str(i),str(j),str(orig),str(dest)), lowBound = 0)
			
# 		# Add decision variable to the dictionary data structure
# 		a_dict['dvFlows'][k] = var
			

# # The objective function is added to 'prob' first
# # lpSum takes a list of coefficients*LpVariables and makes a summation
# prob += pulp.lpSum(objFn), "Total Cost"

pulp.objFn.append(lpSum(y[i]*F for i in Zip), "Facility Build Cost")
for m in M:
	for i in Zip:
		for j in CustLoc:
			prob += pulp.lpSum([C[m][i][j]*d[j][val]*V[val]*x[i,j,m,val] for val in V]), "Cost to transport"
>>>>>>> master

 #objective function
 #V[k]*C[i,j,m]*D[i,k]*x[i,j,m,k]   +F*y[j]
#C= {ori:{dest: ***, dest2:***} ori2: {dest:...}}

for a,a_dict in Combo.iteritems():
	origin=a['route'][1] #zip
	dest=a['route'][0] #customer

prob += pulp.lpSum(objFn), "Total Cost"
#[varAH, varAL, varTH, varTL]
#combo={(1,2):{'customer':1, 'location':2, 'dv':[1,0,0,0]},(1,3):{'customer':1, 'location':3, 'dv':[0,0,0,0]}}
#combo ={(i,j):{'customer':i, 'location':j,'dv':[x1,x2,x3,x4]}}


#constraint 1
for i,j in zip(Zip, CustLoc):
	prob+= shipmax <= shipmax + M*(1-combo[(i,j)]['dv'][0]) #air,high
	prob+= shipmax <=shipmax + M*(1-combo[(i,j)]['dv'][1]) #air, low
	prob += trucktime[i][j] <= shipmax + M*(1-combo[(i,j)]['dv'][2])#truck, high
	prob += trucktime[i][j] <= shipmax + M*(1-combo[(i,j)]['dv'][3]) #truck, low
     #for all i,j,m,k: t[i,j,m]<=1+M*(1-x[i,j,m,k])
     #for all i,j,m,k: t[i,j,m]<=2+M*(1-x[i,j,m,k]) 

<<<<<<< HEAD
#constraint 2
#obj['dv'] for route,obj in combo.items() if obj['customer'] == i
for i in CustLoc:
	low=[]
	high=[]
	for a,a_dict in combo.iterItems():
		if a['customer']==i:
			low += [a['dv'][1], a['dv'][3]]
			high += [a['dv'][0], a['dv'][2]]
	prob+= sum(low) =1
	prob+= sum(high) =1
    
#constraint 3
for i in Zip:
 	prob += LpSum(sum(combo[(i,j)]['dv']) for j in CustLoc) <= M*y[i]
=======
 #constraint 2
for i in Zip:
 	for j in CustLoc:
 		for m in M:
 			prob+= pulp.LpSum(x[i,j,m,k] for k in V) >=1
 		 #for all i,j: sum[m,k](x[i,j,m,k])>=1
 
 #constraint 3
for j in CustLoc:
 	for m in M:
 		for k in V:
 			prob+=pulp.LpSum(x[i,j,m,k] for i in Zip) <=M*y[j]
>>>>>>> master
 #for all j: sum[k,m,i](x[i,j,m,k])<=M[j]

# Write out as a .LP file
prob.writeLP("UPS_network.lp")

# The problem is solved using PuLP's choice of Solver
prob.solve(GUROBI())
#prob.solve()

# The status of the solution is printed to the screen
print ("Status:", pulp.LpStatus[prob.status])

# Each of the variables is printed with it's resolved optimum value
# for a in x:
#     print( v.name, "=", v.varValue)
#print out only locations that were planted. also print out #no.
<<<<<<< HEAD
  
print "Total Cost = ", value(prob.objective)
<<<<<<< HEAD
=======
=======

# The optimised objective function value is printed to the screen    
print ("Total Cost = ", value(prob.objective))
>>>>>>> master

	
 

 
 
""" 
<<<<<<< HEAD
 
=======
 #constraint 4:
 #for all j:
 #x[1,c]+x[2,c]+x[3,c] <= Y[j]
>>>>>>> origin/rachel
<<<<<<< HEAD
>>>>>>> 1779c6a5045247f6f0f855da95b45e8c51cb0d47
=======
"""
>>>>>>> master
