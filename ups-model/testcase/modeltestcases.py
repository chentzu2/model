# -*- coding: utf-8 -*-
"""
Created on Tue Apr  3 21:08:04 2018

@author: Tzu-Ying, Pear, Yi
Project name: UPS Optimization Model for Network Design
"""

#testing github desktop

import pulp
import math
import openpyxl
file = open('test.txt','w')
data = openpyxl.load_workbook('list_zipcode_testcase1.xlsx',read_only=True, data_only=True)
sheet = data['Sheet1']
title = []
for row in sheet.rows:
    
    for cell in row:
        title.append(cell.value)


data = openpyxl.load_workbook('cost_matrix_truck_testcase1.xlsx',read_only=True, data_only=True)
sheet = data['cost']

cost_truck = {}

#inputting cost for the truck 
# structure would be like {ori:{dest: ***, dest2:***} ori2: {dest:...}}
for row in sheet.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    ori = temp[0]
    temp.pop(0)
    cost_truck[ori] = {}
    for i in range(len(title)):
        cost_truck[ori][title[i]] = []
        cost_truck[ori][title[i]].append(temp[i])



data2 = openpyxl.load_workbook('cost_matrix_deliveryday_testcase1.xlsx', read_only=True, data_only=True)
sheet2 = data2['day']

delivery_day = {}

#inputting require days for each ori to dest
# structure would be like {ori:{dest: ***, dest2:***} ori2: {dest:...}}

for row in sheet2.rows:
    temp = []

    for cell in row:
        temp.append(cell.value)
    ori = temp[0]
    temp.pop(0)
    delivery_day[ori] = {}
    for i in range(len(title)):
        delivery_day[ori][title[i]] = []
        delivery_day[ori][title[i]].append(temp[i])



# creating dictionary for next day air cost from ori to dest
# structure would be like {ori:{dest: ***, dest2:***} ori2: {dest:...}}
data = openpyxl.load_workbook('cost_matrix_nda_testcase1.xlsx',read_only=True, data_only=True)
sheet = data['cost']

cost_nda = {}

for row in sheet.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    ori = temp[0]
    temp.pop(0)
    cost_nda[ori] = {}
    for i in range(len(title)):
        cost_nda[ori][title[i]] = []
        cost_nda[ori][title[i]].append(temp[i])


# creating dictionary for 2nd day air cost from ori to dest
# structure would be like {ori:{dest: ***, dest2:***} ori2: {dest:...}}
data = openpyxl.load_workbook('cost_matrix_sda_testcase1.xlsx',read_only=True, data_only=True)
sheet = data['cost']

cost_sda = {}

for row in sheet.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    ori = temp[0]
    temp.pop(0)
    cost_sda[ori] = {}
    for i in range(len(title)):
        cost_sda[ori][title[i]] = []
        cost_sda[ori][title[i]].append(temp[i])

# creating dictionary for demand of high value and value of each 3 digit
# structure would be like {3digit: [high, low], ...}
data = openpyxl.load_workbook('demandtestcase1.xlsx',read_only=True, data_only=True)
sheet = data['Demand']

demand = {}

for row in sheet.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    zipcode = temp[0]
    temp.pop(0)
    demand[zipcode] = []
    for i in range(len(temp)):
        demand[zipcode].append(temp[i])

# creating dictionary for demand of high value and value of each 3 digit
# structure would be like {3digit: cost, 3digit2: cost, ... }
data = openpyxl.load_workbook('facility_cost_testcase1.xlsx',read_only=True, data_only=True)
sheet = data['Demand']

facility_cost = {}

for row in sheet.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    zipcode = temp[0]
    temp.pop(0)
    facility_cost[zipcode] = temp[0]

#pear optimization code

# Data Section
CustLoc=list(demand.keys()) #a set of customer locations
Zip=title # set of 3-digit zip code
shipmax = 1 #change from next day to 2 days - connect to GUI
F =   24243148 #fixed cost to operate/open facility  
travelMode = [1,2] #[air, truck]
M=10000000
costAir={1:cost_nda, 2:cost_sda} #{ori:{dest: ***, dest2:***} ori2: {dest:...}}
trucktime= delivery_day #{} #delivery_day={ori:{dest: ***, dest2:***} ori2: {dest:...}}
costTruck = cost_truck #{ori:{dest: ***, dest2:***} ori2: {dest:...}}
d = demand #{3digit: [high, low], ...}
# Code Section

# Create the 'prob' object to contain the problem data
 #create prob object to contain optimization problem data
prob = pulp.LpProblem("Facility Location Plan", pulp.LpMinimize)

objFn=[]
# Decision variables

combo={} #={(1,2):{'route}}
for i in Zip:
    for j in CustLoc:
        combo[(i,j)]={'location':i, 'customer':j} #origin, destination 


for a,a_dict in combo.items():
    z = a[0]
    cl = a[1]

    varAH = pulp.LpVariable("AirHigh(%s,%s)"%(str(z),str(cl)), lowBound=0, cat='Binary')
    varAL = pulp.LpVariable("AirLow(%s,%s)"%(str(z),str(cl)), lowBound=0, cat='Binary')
    varTH = pulp.LpVariable("TruckHigh(%s,%s)"%(str(z),str(cl)), lowBound=0, cat='Binary')
    varTL = pulp.LpVariable("TruckLow(%s,%s)"%(str(z),str(cl)), lowBound=0, cat='Binary')

    # a_dict['dvAH'] = varAH
    # a_dict['dvAL'] = varAL
    # a_dict['dvTH'] = varTH
    # a_dict['dvTL'] = varTL    
    #value = [l,h]
    a_dict['dv']=[varAH, varAL, varTH, varTL]
    fn=(varAH*float(d[cl][0])*150+varAL*float(d[cl][1])*20)*float(costAir[shipmax][z][cl][0]) + (varTH*float(d[cl][0])*150+varTL*float(d[cl][1])*20)*float(costTruck[z][cl][0])
    print(fn)
    objFn.append(fn)


FacilityLocations={}

for j in Zip:
    dvLoc = pulp.LpVariable("Zip(%s)"%str(j), lowBound=0, upBound=1, cat='Binary')
    FacilityLocations[j] = dvLoc
    objFn.append(dvLoc*F)

prob += pulp.lpSum(objFn), "Total Cost"


 #objective function
 #V[k]*C[i,j,m]*D[i,k]*x[i,j,m,k]   +F*y[j]
#C= {ori:{dest: ***, dest2:***} ori2: {dest:...}}

for a,a_dict in combo.items():
    origin=a[1] #zip
    dest=a[0] #customer

#[varAH, varAL, varTH, varTL]
#combo={(1,2):{'customer':1, 'location':2, 'dv':[1,0,0,0]},(1,3):{'customer':1, 'location':3, 'dv':[0,0,0,0]}}
#combo ={(i,j):{'customer':i, 'location':j,'dv':[x1,x2,x3,x4]}}


#constraint 1
for i in Zip:
    for j in CustLoc:
        prob+= pulp.lpSum(shipmax) <= shipmax + M*(1-combo[(i,j)]['dv'][0]) #air,high
        prob+= pulp.lpSum(shipmax) <=shipmax + M*(1-combo[(i,j)]['dv'][1]) #air, low
        prob += pulp.lpSum(float(trucktime[i][j][0])) <= shipmax + M*(1-combo[(i,j)]['dv'][2])#truck, high
        prob += pulp.lpSum(float(trucktime[i][j][0])) <= shipmax + M*(1-combo[(i,j)]['dv'][3]) #truck, low
         #for all i,j,m,k: t[i,j,m]<=1+M*(1-x[i,j,m,k])
#     #for all i,j,m,k: t[i,j,m]<=2+M*(1-x[i,j,m,k]) 

#all customers with demand fulfilled?
#for j in CustLoc:
#    if sum(d[j])>0:
#        prob+=pulp.lpSum(combo[(i,j)]['dv'] for i in Zip) >=1

#constraint 2
#obj['dv'] for route,obj in combo.items() if obj['customer'] == i
for i in CustLoc:
    low=[]
    high=[]
    for a,a_dict in combo.items():
        if combo[a]['customer']==i:
            low += [combo[a]['dv'][1], combo[a]['dv'][3]]
            high += [combo[a]['dv'][0], combo[a]['dv'][2]]
    prob+= pulp.lpSum(sum(low))==1
    prob+= pulp.lpSum(sum(high))==1

#constraint 3
for i in Zip:
    prob += pulp.lpSum(combo[(i,j)]['dv'] for j in CustLoc) <= M*FacilityLocations[i]
## #for all j: sum[k,m,i](x[i,j,m,k])<=M[j]

# Write out as a .LP file
prob.writeLP("UPS_network.lp")

# The problem is solved using PuLP's choice of Solver
prob.solve(pulp.GUROBI())
#prob.solve()

print ("Status:", pulp.LpStatus[prob.status])
file.write('Status: ')
file.write(str(pulp.LpStatus[prob.status]))
file.write('\n')
for v in prob.variables():
    #file.write(v.name)
    if v.varValue >0:
        file.write(v.name+'='+str(v.varValue) +'\n')
        print(v.name, "=", v.varValue)

#print out only locations that were planted. also print out #no.
  
print ("Total Cost = ", pulp.value(prob.objective))
file.write("Total Cost = ")
file.write(str(pulp.value(prob.objective)))
file.write('\n')
file.close()